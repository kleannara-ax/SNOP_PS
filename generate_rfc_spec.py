#!/usr/bin/env python3
"""
RFC 설계서 Excel 문서 생성
Z_SNOP_PS_OBSOLETE_INV_GET — PS S&OP 진부화재고 리스트 데이터 조회
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 공통 스타일 ──
THIN = Side(style='thin', color='000000')
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

FILL_HEADER = PatternFill('solid', fgColor='1F4E79')       # 진한 네이비
FILL_SUB    = PatternFill('solid', fgColor='D6E4F0')       # 연한 블루
FILL_SECTION = PatternFill('solid', fgColor='2E75B6')      # 중간 블루
FILL_M      = PatternFill('solid', fgColor='E2EFDA')       # 연한 초록
FILL_D      = PatternFill('solid', fgColor='FCE4D6')       # 연한 오렌지
FILL_WARN   = PatternFill('solid', fgColor='FFF2CC')       # 연한 노랑
FILL_WHITE  = PatternFill('solid', fgColor='FFFFFF')
FILL_GRAY   = PatternFill('solid', fgColor='F2F2F2')

FONT_TITLE    = Font(name='맑은 고딕', size=14, bold=True, color='1F4E79')
FONT_HEADER   = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
FONT_SECTION  = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
FONT_SUB      = Font(name='맑은 고딕', size=10, bold=True, color='1F4E79')
FONT_NORMAL   = Font(name='맑은 고딕', size=10)
FONT_BOLD     = Font(name='맑은 고딕', size=10, bold=True)
FONT_SMALL    = Font(name='맑은 고딕', size=9, color='666666')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_TOP    = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_cell(ws, row, col, value='', font=FONT_NORMAL, fill=FILL_WHITE,
               alignment=ALIGN_LEFT, border=BORDER_ALL, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
        for c in range(col + 1, merge_end_col + 1):
            mc = ws.cell(row=row, column=c)
            mc.border = border
            mc.fill = fill
    return cell


def style_header_row(ws, row, col_start, col_end, values,
                     font=FONT_HEADER, fill=FILL_HEADER):
    for i, val in enumerate(values):
        style_cell(ws, row, col_start + i, val, font=font, fill=fill,
                   alignment=ALIGN_CENTER)


def add_label_value_row(ws, row, label, value, col_label=2, col_value=3,
                        col_value_end=8):
    style_cell(ws, row, col_label, label, font=FONT_BOLD, fill=FILL_SUB,
               alignment=ALIGN_CENTER)
    style_cell(ws, row, col_value, value, font=FONT_NORMAL, fill=FILL_WHITE,
               alignment=ALIGN_LEFT, merge_end_col=col_value_end)


# ══════════════════════════════════════════════════════════════
# Sheet 1: RFC 개요
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '1. RFC 개요'
ws1.sheet_properties.tabColor = '1F4E79'

# 컬럼 너비
for col, w in [(1, 3), (2, 18), (3, 20), (4, 20), (5, 15), (6, 15), (7, 15), (8, 25)]:
    ws1.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws1, r, 2, 'RFC 인터페이스 설계서', font=FONT_TITLE,
           fill=FILL_WHITE, alignment=ALIGN_LEFT, merge_end_col=8,
           border=Border())
r += 1
style_cell(ws1, r, 2, 'PS S&OP 진부화재고 리스트 데이터 조회',
           font=Font(name='맑은 고딕', size=11, color='666666'),
           fill=FILL_WHITE, alignment=ALIGN_LEFT, merge_end_col=8,
           border=Border())

# 기본 정보
r = 5
style_cell(ws1, r, 2, '■ 기본 정보', font=FONT_BOLD, fill=FILL_SECTION,
           alignment=ALIGN_LEFT, merge_end_col=8)
ws1.cell(row=r, column=2).font = FONT_SECTION

r += 1; add_label_value_row(ws1, r, 'RFC Function명', 'Z_SNOP_PS_OBSOLETE_INV_GET')
r += 1; add_label_value_row(ws1, r, '설명', 'PS S&OP 진부화재고 리스트 화면에 필요한 데이터를 SAP에서 조회하여 반환')
r += 1; add_label_value_row(ws1, r, '대상 시스템', 'SAP ECC / S4HANA → PS S&OP 웹 시스템')
r += 1; add_label_value_row(ws1, r, '대상 DB 테이블', 'ps_obsolete_inventory (MariaDB)')
r += 1; add_label_value_row(ws1, r, '대상 화면', '진부화재고 리스트 (PS S&OP 계획 시스템)')
r += 1; add_label_value_row(ws1, r, '호출 방향', 'Web Server → SAP (RFC Call)')
r += 1; add_label_value_row(ws1, r, '작성일', '2026-04-30')
r += 1; add_label_value_row(ws1, r, '작성자', '')
r += 1; add_label_value_row(ws1, r, '버전', 'v1.1')

# 호출 유형
r += 2
style_cell(ws1, r, 2, '■ 호출 유형 (IV_SYNC_TYPE)', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=8)

r += 1
style_header_row(ws1, r, 2, 8,
    ['구분', '값', '설명', '호출 주기', '트리거 방식', '대상 필드 범위', '비고'])

rows_type = [
    ['마스터', "'M'", '기준정보 (느리게 변하는 마스터성 데이터)',
     '월 1회 (또는 필요 시)', '사용자 [기준정보 동기화] 버튼 클릭',
     '기초(연수/중량/금액)\n예정(연수/중량/금액)', '월초 또는 기준정보 변경 시'],
    ['일별', "'D'", '재고/실적 (매일 변하는 트랜잭션성 데이터)',
     '매일 1회', '매일 07:00 자동 스케줄\n+ [SAP 연동] 버튼 (수동 보완)',
     '현재고\n매출/밀롤/기타출고 실적', '스케줄 실패 시 수동 버튼으로 보완'],
]
for i, vals in enumerate(rows_type):
    r += 1
    fill = FILL_M if i == 0 else FILL_D
    for j, v in enumerate(vals):
        style_cell(ws1, r, 2 + j, v, font=FONT_NORMAL, fill=fill,
                   alignment=ALIGN_CENTER if j < 2 else ALIGN_LEFT)

# DB 반영 규칙
r += 2
style_cell(ws1, r, 2, '■ DB 반영 규칙', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=8)

r += 1
style_header_row(ws1, r, 2, 8,
    ['호출 유형', 'DB 컬럼', '반영 방식', '설명', '', '', ''])
# Merge empty header cols
ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)

m_rules = [
    ["'M' 마스터", 'base_age / base_weight / base_amount', 'RFC 값으로 UPDATE', '기초 정보 갱신'],
    ["'M' 마스터", 'plan_age / plan_weight / plan_amount', 'RFC 값으로 UPDATE', '예정 정보 갱신'],
    ["'D' 일별", 'current_stock', 'RFC 값으로 UPDATE', '현재고 갱신'],
    ["'D' 일별", 'out_sales_adj', 'ETC_SALES 값으로 리셋', '옵션A: 매일 SAP에서 변동되는 값이므로 리셋'],
    ["'D' 일별", 'out_mill_roll_adj', 'ETC_MILL_ROLL 값으로 리셋', '옵션A: 매일 SAP에서 변동되는 값이므로 리셋'],
    ["'D' 일별", 'out_etc_adj', 'ETC_OUT 값으로 리셋', '옵션A: 매일 SAP에서 변동되는 값이므로 리셋'],
    ["'D' 일별", 'out_disposal', '변경하지 않음 (유지)', '옵션B: SAP 무관 사용자 판단값 — 전일 확정값 유지'],
    ["'D' 일별", 'confirmed_yn', "'N' 으로 리셋", '새 SAP 데이터 → 미확정 상태로 전환'],
    ["'D' 일별", 'confirmed_by / confirmed_dt', 'NULL 로 리셋', '확정자/확정일시 초기화'],
    ["'D' 일별", 'ref_date', 'CURDATE()로 갱신', '기준일자를 당일로 변경'],
    ["'D' 일별", 'last_sync_dt', 'NOW()로 갱신', 'SAP 최종 연동 일시 기록'],
]
for vals in m_rules:
    r += 1
    fill = FILL_M if "'M'" in vals[0] else FILL_D
    style_cell(ws1, r, 2, vals[0], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws1, r, 3, vals[1], font=Font(name='맑은 고딕', size=10, bold=True), fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws1, r, 4, vals[2], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws1, r, 5, vals[3], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT, merge_end_col=8)


# ══════════════════════════════════════════════════════════════
# Sheet 2: Input Parameters
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('2. Input Parameters')
ws2.sheet_properties.tabColor = '2E75B6'

for col, w in [(1, 3), (2, 5), (3, 22), (4, 15), (5, 12), (6, 10), (7, 10), (8, 35), (9, 30)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws2, r, 2, '■ Input Parameters (IMPORTING)', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=9)

r += 1
style_header_row(ws2, r, 2, 9,
    ['No', '파라미터명', 'ABAP Type', 'Length', '필수여부', 'Default', '설명', '비고'])

params = [
    ['1', 'IV_SYNC_TYPE', 'CHAR', '1', '필수', '-',
     "호출 유형\n'M' = 마스터 (기준정보)\n'D' = 일별 (재고/실적)",
     "'M' 또는 'D' 외의 값이 들어오면\nEV_MESSAGE에 오류 반환"],
    ['2', 'IV_PLANT_CODE', 'CHAR', '20', '선택', "'' (빈값)",
     '플랜트 코드 필터\n빈값 = 전체 플랜트 조회',
     '특정 플랜트만 조회할 때 사용\n예: PS10, PS20'],
    ['3', 'IV_MATERIAL', 'CHAR', '50', '선택', "'' (빈값)",
     '자재코드 필터\n빈값 = 전체 자재 조회',
     '특정 자재만 조회할 때 사용'],
]
for vals in params:
    r += 1
    fill = FILL_WHITE if int(vals[0]) % 2 == 1 else FILL_GRAY
    for j, v in enumerate(vals):
        al = ALIGN_CENTER if j in [0, 3, 4, 5] else ALIGN_LEFT
        f = FONT_BOLD if j == 1 else FONT_NORMAL
        style_cell(ws2, r, 2 + j, v, font=f, fill=fill, alignment=al)


# ══════════════════════════════════════════════════════════════
# Sheet 3: Output Parameters
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('3. Output Parameters')
ws3.sheet_properties.tabColor = '2E75B6'

for col, w in [(1, 3), (2, 5), (3, 22), (4, 15), (5, 12), (6, 35), (7, 35)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws3, r, 2, '■ Output Parameters (EXPORTING)', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=7)

r += 1
style_header_row(ws3, r, 2, 7,
    ['No', '파라미터명', 'ABAP Type', 'Length', '설명', '비고'])

out_params = [
    ['1', 'EV_RESULT_CNT', 'INT', '10', '조회 결과 건수', '0이면 해당 조건의 데이터 없음'],
    ['2', 'EV_MESSAGE', 'STRING', '255',
     '처리 결과 메시지\nS: 성공 / E: 오류 / W: 경고',
     "예시:\n'S: 150건 조회 완료'\n'E: IV_SYNC_TYPE 값 오류'\n'W: 일부 자재 누락 (3건)'"],
]
for vals in out_params:
    r += 1
    fill = FILL_WHITE if int(vals[0]) % 2 == 1 else FILL_GRAY
    for j, v in enumerate(vals):
        al = ALIGN_CENTER if j in [0, 3] else ALIGN_LEFT
        f = FONT_BOLD if j == 1 else FONT_NORMAL
        style_cell(ws3, r, 2 + j, v, font=f, fill=fill, alignment=al)


# ══════════════════════════════════════════════════════════════
# Sheet 4: Output Table (ET_DATA)
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('4. Output Table (ET_DATA)')
ws4.sheet_properties.tabColor = '2E75B6'

for col, w in [(1, 3), (2, 5), (3, 22), (4, 18), (5, 12), (6, 10),
               (7, 10), (8, 30), (9, 18), (10, 30)]:
    ws4.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws4, r, 2, '■ Output Table: ET_DATA (TABLES)', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=10)

r += 1
style_header_row(ws4, r, 2, 10,
    ['No', '필드명', 'ABAP Type', 'Length', '소수점',
     "'M' 반환", "'D' 반환", '설명', 'SAP 소스 참고'])

fields = [
    ['1', 'PLANT_CODE', 'CHAR', '20', '-', 'O', 'O',
     '플랜트 코드', 'T001W-WERKS'],
    ['2', 'MATERIAL_CODE', 'CHAR', '50', '-', 'O', 'O',
     '자재코드', 'MARA-MATNR'],
    ['3', 'BASE_AGE', 'INT', '4', '-', 'O', '-',
     '기초 - 연수', '(산출 로직에 따라)'],
    ['4', 'BASE_WEIGHT', 'DEC', '15', '2', 'O', '-',
     '기초 - 중량(톤)', '(산출 로직에 따라)'],
    ['5', 'BASE_AMOUNT', 'DEC', '15', '2', 'O', '-',
     '기초 - 금액(백만원)', '(산출 로직에 따라)'],
    ['6', 'PLAN_AGE', 'INT', '4', '-', 'O', '-',
     '예정 - 연수', '(산출 로직에 따라)'],
    ['7', 'PLAN_WEIGHT', 'DEC', '15', '2', 'O', '-',
     '예정 - 중량(톤)', '(산출 로직에 따라)'],
    ['8', 'PLAN_AMOUNT', 'DEC', '15', '2', 'O', '-',
     '예정 - 금액(백만원)', '(산출 로직에 따라)'],
    ['9', 'CURRENT_STOCK', 'DEC', '15', '2', '-', 'O',
     '현재고', 'MARD-LABST 등'],
    ['10', 'ETC_SALES', 'DEC', '15', '2', '-', 'O',
     '매출 실적', 'VBRP 등 매출 관련'],
    ['11', 'ETC_MILL_ROLL', 'DEC', '15', '2', '-', 'O',
     '밀롤 실적', '(사내 이동 등)'],
    ['12', 'ETC_OUT', 'DEC', '15', '2', '-', 'O',
     '기타출고 실적', 'MSEG 등 출고 관련'],
]

for vals in fields:
    r += 1
    is_m = vals[5] == 'O'
    is_d = vals[6] == 'O'
    if is_m and is_d:
        fill = FILL_WHITE
    elif is_m:
        fill = FILL_M
    else:
        fill = FILL_D
    for j, v in enumerate(vals):
        al = ALIGN_CENTER if j in [0, 3, 4, 5, 6] else ALIGN_LEFT
        f = FONT_BOLD if j == 1 else FONT_NORMAL
        style_cell(ws4, r, 2 + j, v, font=f, fill=fill, alignment=al)

# 범례
r += 2
style_cell(ws4, r, 2, '※ 범례', font=FONT_BOLD, fill=FILL_WHITE,
           alignment=ALIGN_LEFT, merge_end_col=10, border=Border())
r += 1
style_cell(ws4, r, 2, '', font=FONT_NORMAL, fill=FILL_M,
           alignment=ALIGN_CENTER)
style_cell(ws4, r, 3, "O = 해당 호출 유형에서 값을 반환하는 필드  /  - = 빈값(0 또는 공백)으로 반환",
           font=FONT_SMALL, fill=FILL_WHITE, alignment=ALIGN_LEFT,
           merge_end_col=10, border=Border())
r += 1
style_cell(ws4, r, 2, '', font=FONT_NORMAL, fill=FILL_M,
           alignment=ALIGN_CENTER)
style_cell(ws4, r, 3, "초록 배경 = 'M'(마스터) 전용 필드",
           font=FONT_SMALL, fill=FILL_M, alignment=ALIGN_LEFT,
           merge_end_col=5, border=Border())
style_cell(ws4, r, 6, '', font=FONT_NORMAL, fill=FILL_D,
           alignment=ALIGN_CENTER)
style_cell(ws4, r, 7, "주황 배경 = 'D'(일별) 전용 필드",
           font=FONT_SMALL, fill=FILL_D, alignment=ALIGN_LEFT,
           merge_end_col=10, border=Border())


# ══════════════════════════════════════════════════════════════
# Sheet 5: 호출 시나리오 + 에러 처리
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('5. 호출 시나리오')
ws5.sheet_properties.tabColor = '548235'

for col, w in [(1, 3), (2, 5), (3, 22), (4, 20), (5, 15), (6, 35), (7, 30)]:
    ws5.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws5, r, 2, '■ 호출 시나리오', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=7)

r += 1
style_header_row(ws5, r, 2, 7,
    ['No', '시나리오', '호출 유형', '트리거', '상세 설명', '비고'])

scenarios = [
    ['1', '월초 기준정보 갱신',  "'M'", '사용자 버튼 클릭\n[기준정보 동기화]',
     '기초/예정 연수·중량·금액을\nSAP 최신 값으로 갱신\n\n기존 사용자 수정값(out_*)은\n영향 없음\n\n[버튼 호출 흐름]\n1. 사용자 [기준정보 동기화] 클릭\n2. 확인 팝업 표시\n3. RFC M 호출 → 결과 수신\n4. 건수 + 동기화 시각 표시\n5. 버튼 옆 최종 동기화 일시 갱신',
     '매월 초 1회 권장\n기준정보 변경 시 수시 가능\n\n중복 클릭 방지:\n호출 중 버튼 비활성화\n\n최종 동기화 일시 표시:\n버튼 옆에 항상 표시'],
    ['2', '일일 재고/실적 갱신', "'D'", '매일 07:00 자동 스케줄',
     '현재고, 매출/밀롤/기타출고 실적을\nSAP 최신 값으로 갱신\n\n옵션A (매일 SAP 변동):\n  out_sales_adj ← ETC_SALES\n  out_mill_roll_adj ← ETC_MILL_ROLL\n  out_etc_adj ← ETC_OUT\n\n옵션B (사용자 판단값 유지):\n  out_disposal → 전일 확정값 유지\n\nconfirmed_yn → N 리셋',
     '스케줄 실패 시\n사용자 [SAP 연동] 버튼으로\n수동 보완 가능'],
    ['3', '수동 재고/실적 갱신', "'D'", '사용자 버튼 클릭\n[SAP 연동]',
     '시나리오 2와 동일한 처리\n\n긴급 데이터 갱신이 필요하거나\n자동 스케줄 실패 시 사용',
     '담당자 이상 권한 필요'],
    ['4', '전체 갱신 (초기 세팅)', "'M' → 'D'\n순차 호출", '관리자 수동',
     'M 호출로 기준정보 갱신 후\nD 호출로 재고/실적 갱신\n\n시스템 최초 오픈 또는\n데이터 전체 리프레시 시 사용',
     '별도 RFC 불필요\n웹 서버에서 순차 호출로 처리'],
]
for vals in scenarios:
    r += 1
    fill = FILL_WHITE if int(vals[0]) % 2 == 1 else FILL_GRAY
    for j, v in enumerate(vals):
        al = ALIGN_CENTER if j in [0, 2] else ALIGN_LEFT
        style_cell(ws5, r, 2 + j, v, font=FONT_NORMAL, fill=fill, alignment=al)

# 에러 처리
r += 2
style_cell(ws5, r, 2, '■ 에러 처리 / 예외 상황', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=7)

r += 1
style_header_row(ws5, r, 2, 7,
    ['No', '상황', 'EV_MESSAGE', '처리 방법', '비고', ''])
ws5.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

errors = [
    ['1', "IV_SYNC_TYPE에 'M','D' 외의\n값이 입력된 경우",
     "E: IV_SYNC_TYPE 값 오류\n('M' 또는 'D'만 허용)",
     'ET_DATA 빈 테이블 반환\nEV_RESULT_CNT = 0', ''],
    ['2', '해당 플랜트/자재가\nSAP에 존재하지 않는 경우',
     "W: 조회 결과 없음\n(PLANT_CODE=PS99)",
     'ET_DATA 빈 테이블 반환\nEV_RESULT_CNT = 0', '오류가 아닌 경고(W) 처리'],
    ['3', 'SAP 내부 오류\n(DB Lock, 권한 등)',
     "E: SAP 내부 오류 (상세메시지)",
     'ET_DATA 빈 테이블 반환\nEV_RESULT_CNT = 0\n웹 서버에서 재시도 또는 알림', ''],
    ['4', '조회 데이터 건수가\n매우 많은 경우 (10,000건 이상)',
     "S: 15,000건 조회 완료\n(대량 데이터 주의)",
     '정상 반환하되 메시지에\n대량 데이터 경고 포함', 'IV_PLANT_CODE로\n범위 제한 권장'],
]
for vals in errors:
    r += 1
    fill = FILL_WARN if vals[0] in ['1', '3'] else FILL_WHITE
    for j, v in enumerate(vals):
        al = ALIGN_CENTER if j == 0 else ALIGN_LEFT
        style_cell(ws5, r, 2 + j, v, font=FONT_NORMAL, fill=fill, alignment=al)
        if j == 4:
            ws5.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)


# ══════════════════════════════════════════════════════════════
# Sheet 6: 운영 가이드
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('6. 운영 가이드')
ws6.sheet_properties.tabColor = '548235'

for col, w in [(1, 3), (2, 18), (3, 25), (4, 45), (5, 30)]:
    ws6.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws6, r, 2, '■ 운영 환경 설정', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=5)

r += 1
style_header_row(ws6, r, 2, 5, ['항목', '설정값', '설명', '비고'])

ops = [
    ['SAP 쪽 타임아웃', '30초', 'RFC Function Module 실행 제한 시간', '데이터 양에 따라 조정'],
    ['웹 서버 쪽 타임아웃', '60초', 'RFC 호출 대기 제한 시간', 'SAP보다 넉넉하게 설정'],
    ['일별 스케줄 시간', '매일 07:00', "'D' 호출 자동 실행 시간", '업무 시작 전 완료되도록'],
    ['재시도 횟수', '최대 3회', '스케줄 실패 시 자동 재시도', '5분 간격으로 재시도 권장'],
    ['동시 호출 제한', '1회', '같은 SYNC_TYPE 중복 호출 방지', '버튼 클릭 시 더블클릭 방지 필수'],
]
for i, vals in enumerate(ops):
    r += 1
    fill = FILL_WHITE if i % 2 == 0 else FILL_GRAY
    for j, v in enumerate(vals):
        style_cell(ws6, r, 2 + j, v, font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)

# 권한
r += 2
style_cell(ws6, r, 2, '■ 호출 권한', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=5)

r += 1
style_header_row(ws6, r, 2, 5, ['호출 유형', '허용 주체', '설명', ''])
ws6.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

auths = [
    ["'M' 수동 버튼", '담당자 (PLANNER) 이상', '기준정보 동기화 버튼은 담당자 이상만 클릭 가능'],
    ["'D' 자동 스케줄", '시스템 계정 (SAP_INTERFACE)', '매일 07:00 자동 실행 — 시스템 계정으로 호출'],
    ["'D' 수동 버튼", '담당자 (PLANNER) 이상', 'SAP 연동 버튼은 담당자 이상만 클릭 가능'],
]
for i, vals in enumerate(auths):
    r += 1
    fill = FILL_WHITE if i % 2 == 0 else FILL_GRAY
    style_cell(ws6, r, 2, vals[0], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws6, r, 3, vals[1], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws6, r, 4, vals[2], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT, merge_end_col=5)

# 모니터링
r += 2
style_cell(ws6, r, 2, '■ 모니터링 항목', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=5)

r += 1
style_header_row(ws6, r, 2, 5, ['항목', '확인 위치', '확인 내용', ''])
ws6.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

monitors = [
    ['최종 동기화 일시', '진부화재고 리스트 화면', "last_sync_dt 값 — 당일 07:00 이후인지 확인"],
    ['확정 상태', '진부화재고 리스트 화면', "confirmed_yn — 'D' 호출 후 'N'으로 리셋되었는지 확인"],
    ['동기화 건수', 'EV_RESULT_CNT / 로그', '조회 건수가 예상 범위인지 확인 (급감/급증 이상징후)'],
    ['오류 메시지', 'EV_MESSAGE / 로그', "'E:' 시작 메시지가 있으면 즉시 확인"],
]
for i, vals in enumerate(monitors):
    r += 1
    fill = FILL_WHITE if i % 2 == 0 else FILL_GRAY
    style_cell(ws6, r, 2, vals[0], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws6, r, 3, vals[1], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws6, r, 4, vals[2], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT, merge_end_col=5)


# ══════════════════════════════════════════════════════════════
# Sheet 7: 필드별 리셋/유지 규칙 상세
# ══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet('7. 필드별 리셋·유지 규칙')
ws7.sheet_properties.tabColor = 'C00000'

for col, w in [(1, 3), (2, 5), (3, 22), (4, 18), (5, 12), (6, 20), (7, 45)]:
    ws7.column_dimensions[get_column_letter(col)].width = w

r = 2
style_cell(ws7, r, 2, '■ SAP 일별 동기화(D) 시 사용자 수정 필드 처리 규칙', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=7)

r += 1
style_cell(ws7, r, 2, '확정 일자: 2026-04-30  |  v1.1 신규 추가',
           font=Font(name='맑은 고딕', size=9, color='666666', italic=True),
           fill=FILL_WHITE, alignment=ALIGN_LEFT, merge_end_col=7, border=Border())

r += 2
style_header_row(ws7, r, 2, 7,
    ['No', 'DB 컬럼명', '화면 표시명', '옵션', '동기화 시 처리', '사유 및 상세'])

FILL_RESET = PatternFill('solid', fgColor='FCE4D6')   # 연한 오렌지 (리셋)
FILL_KEEP  = PatternFill('solid', fgColor='E2EFDA')    # 연한 초록 (유지)

field_rules = [
    ['1', 'out_sales_adj', '매출 (수정)', '옵션A',
     'SAP 값으로 리셋',
     '매일 SAP에서 변동되는 매출 실적값\n→ 동기화 시 SAP RFC \'D\' 호출 결과(ETC_SALES)로 덮어씌움\n→ 사용자가 전일 입력한 값은 초기화됨'],
    ['2', 'out_mill_roll_adj', '밀롤 (수정)', '옵션A',
     'SAP 값으로 리셋',
     '매일 SAP에서 변동되는 밀롤 실적값\n→ 동기화 시 SAP RFC \'D\' 호출 결과(ETC_MILL_ROLL)로 덮어씌움\n→ 사용자가 전일 입력한 값은 초기화됨'],
    ['3', 'out_etc_adj', '기타출고 (수정)', '옵션A',
     'SAP 값으로 리셋',
     '매일 SAP에서 변동되는 기타출고 실적값\n→ 동기화 시 SAP RFC \'D\' 호출 결과(ETC_OUT)로 덮어씌움\n→ 사용자가 전일 입력한 값은 초기화됨'],
    ['4', 'out_disposal', '폐기', '옵션B',
     '전일 확정값 유지',
     'SAP와 무관한 사용자 판단값\n→ 동기화 시 변경하지 않음 (전일 사용자가 입력/확정한 값 그대로)\n→ 폐기 수량은 SAP 실적이 아닌 사용자 계획치'],
]

for vals in field_rules:
    r += 1
    fill = FILL_RESET if vals[3] == '옵션A' else FILL_KEEP
    style_cell(ws7, r, 2, vals[0], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws7, r, 3, vals[1], font=FONT_BOLD, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws7, r, 4, vals[2], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws7, r, 5, vals[3], font=FONT_BOLD, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws7, r, 6, vals[4], font=FONT_BOLD, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws7, r, 7, vals[5], font=FONT_NORMAL, fill=fill, alignment=ALIGN_TOP)

# 공통 처리 필드
r += 2
style_cell(ws7, r, 2, '■ 공통 처리 필드 (모든 동기화 시)', font=FONT_SECTION,
           fill=FILL_SECTION, alignment=ALIGN_LEFT, merge_end_col=7)

r += 1
style_header_row(ws7, r, 2, 7,
    ['No', 'DB 컬럼명', '처리 내용', '', '', '비고'])
ws7.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

common_fields = [
    ['1', 'confirmed_yn', "'N' 으로 리셋 (미확정 전환)", '새 SAP 데이터 수신 → 재확정 필요'],
    ['2', 'confirmed_by', 'NULL 으로 리셋', '확정자 정보 초기화'],
    ['3', 'confirmed_dt', 'NULL 으로 리셋', '확정일시 초기화'],
    ['4', 'ref_date', 'CURDATE() 로 갱신', '기준일자를 동기화 당일로 변경'],
    ['5', 'last_sync_dt', 'NOW() 로 갱신', 'SAP 최종 연동 일시 기록'],
]

for i, vals in enumerate(common_fields):
    r += 1
    fill = FILL_WHITE if i % 2 == 0 else FILL_GRAY
    style_cell(ws7, r, 2, vals[0], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    style_cell(ws7, r, 3, vals[1], font=FONT_BOLD, fill=fill, alignment=ALIGN_LEFT)
    style_cell(ws7, r, 4, vals[2], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT, merge_end_col=6)
    style_cell(ws7, r, 7, vals[3], font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT)

# 범례
r += 2
style_cell(ws7, r, 2, '※ 범례', font=FONT_BOLD, fill=FILL_WHITE,
           alignment=ALIGN_LEFT, merge_end_col=7, border=Border())
r += 1
style_cell(ws7, r, 2, '', font=FONT_NORMAL, fill=FILL_RESET, alignment=ALIGN_CENTER)
style_cell(ws7, r, 3, '옵션A (리셋) — 매일 SAP에서 변동되는 값이므로 동기화 시 SAP 값으로 리셋',
           font=FONT_SMALL, fill=FILL_RESET, alignment=ALIGN_LEFT, merge_end_col=7, border=Border())
r += 1
style_cell(ws7, r, 2, '', font=FONT_NORMAL, fill=FILL_KEEP, alignment=ALIGN_CENTER)
style_cell(ws7, r, 3, '옵션B (유지) — SAP와 무관한 사용자 판단값이므로 전일 확정값 그대로 유지',
           font=FONT_SMALL, fill=FILL_KEEP, alignment=ALIGN_LEFT, merge_end_col=7, border=Border())


# ── 인쇄 설정 ──
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

# ── 저장 ──
output_path = '/home/user/webapp/docs/RFC_설계서_Z_SNOP_PS_OBSOLETE_INV_GET_v1.1.xlsx'
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f'RFC 설계서 생성 완료: {output_path}')
